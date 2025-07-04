
import os
import json
import hashlib
from datetime import datetime
from pymongo import MongoClient
from PyPDF2 import PdfReader
import openpyxl

# Configurações
JSON_MODELO = "/home/neirivon/backup_dados_pesados/PISA_novo/PROTOCOLOS/2022/modelo_extracao_pisa_2022.json"
LOG_ARQUIVO = "/home/neirivon/backup_dados_pesados/PISA_novo/PROTOCOLOS/2022/log_extracao_pisa_2022.txt"
MONGO_URI = "mongodb://admin:admin123@localhost:27017/?authSource=admin"
DB_NAME = "pisa"

def gerar_hash_binario(caminho_arquivo):
    with open(caminho_arquivo, "rb") as f:
        conteudo_binario = f.read()
    return hashlib.sha256(conteudo_binario).hexdigest()

def extrair_texto_pdf(caminho, inicio, fim):
    try:
        reader = PdfReader(caminho)
        texto = ""
        for i in range(inicio - 1, fim):
            texto += reader.pages[i].extract_text() or ""
        return texto.strip()
    except Exception as e:
        return f"[ERRO AO EXTRAIR TEXTO PDF] {e}"

def extrair_texto_xlsx(caminho):
    try:
        wb = openpyxl.load_workbook(caminho, read_only=True)
        texto = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                linha = " | ".join([str(cell) if cell is not None else "" for cell in row])
                texto += linha + "\n"
        return texto.strip()
    except Exception as e:
        return f"[ERRO AO EXTRAIR XLSX] {e}"

def extrair_trecho_representativo(texto_completo):
    for paragrafo in texto_completo.split("\n"):
        if len(paragrafo.strip()) > 80:
            return paragrafo.strip()
    return texto_completo.strip().split("\n")[0]

def log(mensagem):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    linha = f"[{timestamp}] {mensagem}"
    print(linha)
    with open(LOG_ARQUIVO, "a", encoding="utf-8") as f:
        f.write(linha + "\n")

def executar_extracao():
    log("🚀 Iniciando extração PISA 2022...")

    cliente = MongoClient(MONGO_URI)
    db = cliente[DB_NAME]

    with open(JSON_MODELO, "r", encoding="utf-8") as f:
        dados = json.load(f)

    for item in dados:
        caminho = item["fonte"]
        colecao = item["colecao_mongodb"]
        pagina_inicio = item["pagina_inicio"]
        pagina_fim = item["pagina_fim"]

        if not os.path.exists(caminho):
            log(f"❌ Arquivo não encontrado: {caminho}")
            continue

        try:
            hash_arquivo = gerar_hash_binario(caminho)
        except Exception as e:
            log(f"❌ Erro ao gerar hash do arquivo: {caminho} — {e}")
            continue

        if caminho.endswith(".pdf") and isinstance(pagina_inicio, int) and isinstance(pagina_fim, int):
            conteudo_completo = extrair_texto_pdf(caminho, pagina_inicio, pagina_fim)
        elif caminho.endswith(".xlsx"):
            conteudo_completo = extrair_texto_xlsx(caminho)
        else:
            log(f"⚠️ Tipo de arquivo ou páginas inválidas: {caminho}")
            continue

        trecho = extrair_trecho_representativo(conteudo_completo)

        doc = {
            "arquivo_original": os.path.basename(caminho),
            "colecao_mongodb": colecao,
            "pagina_inicio": pagina_inicio,
            "pagina_fim": pagina_fim,
            "trecho_do_conteudo": trecho,
            "hash_conteudo": hash_arquivo,
            "data_extracao": datetime.now().isoformat(),
            "fonte": caminho
        }

        try:
            db[colecao].insert_one(doc)
            log(f"✅ Inserido com sucesso em '{colecao}': {os.path.basename(caminho)}")
        except Exception as e:
            log(f"❌ Erro ao inserir em {colecao}: {e}")

    log("🏁 Extração finalizada.")
    cliente.close()

if __name__ == "__main__":
    executar_extracao()
